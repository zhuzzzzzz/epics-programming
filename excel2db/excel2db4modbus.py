import copy
import os.path
import sys

from openpyxl import load_workbook

drvUser2DTYP = {
    'BCD': 'asynInt32',
    'UINT16': 'asynUInt32Digital',
    'INT16': 'asynInt32',
    'UINT32': 'asynInt32',
    'INT32': 'asynInt32',
    'UINT64': 'asynInt64',
    'INT64': 'asynInt64',
    'FLOAT32': 'asynFloat64',
    'FLOAT64': 'asynFloat64',
    'STRING': 'asynInt32',
    'ZSTRING': 'asynInt32',
}


class ModbusDevice:
    def __init__(self):
        self.name = None
        self.device_type = None
        self.address = None  # ip:port
        self.info = None

    def __repr__(self):
        return f'"{self.name}, {self.device_type}, {self.address}"'


DeviceRegistered = {}


class ModbusRecord:
    def __init__(self, record_name, record_type, record_scan='Passive', **kwargs):
        self.name = record_name
        self.name_prefix = None
        self.name_seperator = '-'
        self.type = record_type  # 'ai' 'ao' 'bi' 'bo'
        self.scan = record_scan
        self.desc = None
        self.prec = None
        self.egu = None
        self.other_fields = []
        #
        self.device = None  #
        self.memory_address = None  # memory address for data in PLC
        self.memory_length = None  # data memory_length
        self.device_access = None  # 'r', 'w' or 'rw'
        self.drvUser_prefix = None
        self.drvUser_suffix = None
        self.modbus_funcode = None  # 3, 16
        self.interface_name = None
        self.memory_address_mask = None

        for key, value in kwargs.items():
            if hasattr(self, key):
                setattr(self, key, value)

    def __str__(self):
        self.gen_prepare()
        res = ''
        for line in self.gen_config_lines():
            res += line
        for line in self.gen_db_lines():
            res += line
        return res

    @staticmethod
    def get_field_line(field_str):
        return f'field({field_str})'

    def gen_prepare(self):
        #
        if not self.name_prefix:
            self.name_prefix = ''
        else:
            self.name_prefix = self.name_prefix + self.name_seperator
        #
        if not self.drvUser_suffix:
            self.drvUser_suffix = ''
        #
        if self.memory_address_mask:
            self.type = 'b'
        else:
            self.type = 'a'
        #
        if self.device_access == 'r':
            self.type = self.type + 'i'
            self.modbus_funcode = 3
            self.interface_name = self.name + f"{self.modbus_funcode}R"
        elif self.device_access == 'w':
            self.type = self.type + 'o'
            self.modbus_funcode = 16
            self.interface_name = self.name + f"{self.modbus_funcode}W"

    def gen_config_lines(self):
        # drvModbusAsynConfigure(portName,
        #                        tcpPortName,
        #                        slaveAddress,
        #                        modbusFunction,
        #                        modbusStartAddress,
        #                        modbusLength,
        #                        dataType,
        #                        pollMsec,
        #                        plcType);
        return (f'drvModbusAsynConfigure('
                f'"{self.interface_name}", '
                f'"{self.device.name}", '
                f'0, '
                f'{self.modbus_funcode}, '
                f'{self.memory_address}, '
                f'{self.memory_length}, '
                f'0, '
                f'100, '
                f'"{self.device.name}")\n')

    def gen_db_lines(self):
        db_lines = [
            f'record({self.type}, "{self.name_prefix + self.name}"){{\n',
        ]
        # DTYP
        line = self.get_field_line(f'DTYP, "{drvUser2DTYP[self.drvUser_prefix]}"')
        db_lines.append(f'\t{line}\n')
        # INP or OUT
        if self.device_access == 'r' and 'a' in self.type:
            line = self.get_field_line(
                f'INP, "@asyn({self.interface_name} 0){self.drvUser_prefix + self.drvUser_suffix}"')
        elif self.device_access == 'r' and 'b' in self.type:
            line = self.get_field_line(
                f'INP, "@asynMask({self.interface_name} 0 {self.memory_address_mask}){self.drvUser_prefix + self.drvUser_suffix}"')
        elif self.device_access == 'w' and 'a' in self.type:
            line = self.get_field_line(
                f'OUT, "@asyn({self.interface_name} 0){self.drvUser_prefix + self.drvUser_suffix}"')
        elif self.device_access == 'w' and 'b' in self.type:
            line = self.get_field_line(
                f'OUT, "@asynMask({self.interface_name} 0 {self.memory_address_mask}){self.drvUser_prefix + self.drvUser_suffix}"')
        else:
            print(f'{self.name}.gen_db_lines failed for line "INP" or "OUT". {self.device_access} {self.type}')
        db_lines.append(f'\t{line}\n')
        # SCAN
        if self.scan:
            line = self.get_field_line(f'SCAN, "{self.scan}"')
            db_lines.append(f'\t{line}\n')
        # DESC
        if self.desc:
            line = self.get_field_line(f'DESC, "{self.desc}"')
            db_lines.append(f'\t{line}\n')
        # PREC
        if self.prec:
            line = self.get_field_line(f'PREC, "{self.prec}"')
            db_lines.append(f'\t{line}\n')
        # EGU
        if self.egu:
            line = self.get_field_line(f'EGU, "{self.egu}"')
            db_lines.append(f'\t{line}\n')
        # other other_fields
        if self.other_fields:
            for item in self.other_fields:
                if item:
                    line = self.get_field_line(item)
                    db_lines.append(f'\t{line}\n')
        # end
        db_lines.append(f'}}\n')
        return db_lines


# return a list of EXCEL contents, merged_cells are handled.
def get_excel_cells(file_path='./modbus2db.xlsx', sheet_name='example'):
    workbook = load_workbook(file_path)
    sheet_loaded = workbook[sheet_name]
    # print(sheet_loaded.calculate_dimension())

    excel_list = []
    for row in sheet_loaded.iter_rows():
        line_list = []
        for cell in row:
            line_list.append(cell.value)
        else:
            excel_list.append(line_list)

    for merged_cell in sheet_loaded.merged_cells:
        merged_value = sheet_loaded.cell(row=merged_cell.min_row, column=merged_cell.min_col).value
        for i in range(merged_cell.bounds[0] - 1, merged_cell.bounds[2]):
            for j in range(merged_cell.bounds[1] - 1, merged_cell.bounds[3]):
                excel_list[j][i] = merged_value

    # for line_list in excel_list:
    #     for item in line_list:
    #         print(item, end=' ')
    #     else:
    #         print('')

    return excel_list


def get_pv_info(excel_list):
    pv_title = excel_list[0]
    pv_list = []
    for j in range(1, len(excel_list)):
        pv_temp = ModbusRecord(record_name=None, record_type=None)
        #
        device_name = None
        device_address = None
        #
        for i in range(len(pv_title)):
            if 'PLC名称' in pv_title[i]:
                device_name = excel_list[j][i]
            elif 'IP:Port' in pv_title[i]:
                device_address = excel_list[j][i]
            elif 'Address' in pv_title[i]:
                pv_temp.memory_address = excel_list[j][i]
            elif '数据操作' in pv_title[i]:
                pv_temp.device_access = excel_list[j][i]
            elif '数据长度' in pv_title[i]:
                pv_temp.memory_length = excel_list[j][i]
            elif 'PV前缀' in pv_title[i]:
                pv_temp.name_prefix = excel_list[j][i]
            elif 'PV后缀' in pv_title[i]:
                pv_temp.name = excel_list[j][i]
            elif '更新周期' in pv_title[i]:
                pv_temp.scan = excel_list[j][i]
            elif 'PV描述' in pv_title[i]:
                pv_temp.desc = excel_list[j][i]
            elif '数据精度' in pv_title[i]:
                pv_temp.prec = excel_list[j][i]
            elif '数据单位' in pv_title[i]:
                pv_temp.egu = excel_list[j][i]
            elif '数据类型' in pv_title[i]:
                pv_temp.drvUser_prefix = excel_list[j][i]
            elif '数据格式' in pv_title[i]:
                pv_temp.drvUser_suffix = excel_list[j][i]
            elif '掩码' in pv_title[i]:
                pv_temp.memory_address_mask = excel_list[j][i]
            elif '其他EPICS字段' in pv_title[i]:
                pv_temp.other_fields.append(excel_list[j][i])
        #
        if device_name and device_name not in DeviceRegistered.keys():
            DeviceRegistered[device_name] = ModbusDevice()
            DeviceRegistered[device_name].name = device_name
            DeviceRegistered[device_name].address = device_address
            DeviceRegistered[device_name].device_type = 'Modbus'
        if device_name:
            pv_temp.device = DeviceRegistered[device_name]
        #
        if pv_temp.device_access == 'rw':
            pv_temp_r = copy.deepcopy(pv_temp)
            pv_temp_r.device_access = 'r'
            pv_temp_r.name += 'R'
            pv_temp_w = copy.deepcopy(pv_temp)
            pv_temp_w.device_access = 'w'
            pv_temp_w.name += 'W'
            pv_list.append(pv_temp_r)
            pv_list.append(pv_temp_w)
        else:
            pv_list.append(pv_temp)
    return pv_list


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(f'{sys.argv[0]}: At least one argument should be specified!\n'
              f'arg: "excel path"\n'
              f'arg1: "excel path" \targ2: "object file"\n'
              f'arg1: "excel path" \targ2: "sheet name" \targ3:"object file"\n')
    else:
        if not os.path.isfile(sys.argv[1]):
            print(f'{sys.argv[0]}: Invalid excel path: {sys.argv[1]}!')
            exit(1)
        file_name = ''
        pl = None
        if len(sys.argv) == 2:
            if os.path.isfile(sys.argv[1]):
                pl = get_pv_info(get_excel_cells(file_path=sys.argv[1]))
                for key in DeviceRegistered.keys():
                    file_name += key
                # for item in pl:
                #     print(item)
                # else:
                #     print('')
        elif len(sys.argv) == 3:
            pl = get_pv_info(get_excel_cells(file_path=sys.argv[1]))
            file_name = sys.argv[2]
        else:
            pl = get_pv_info(get_excel_cells(file_path=sys.argv[1], sheet_name=sys.argv[2]))
            file_name = sys.argv[3]
        # generate db files for modbus
        #
        object_path = os.path.join(os.getcwd(), 'res')
        try:
            os.makedirs(object_path)
        except Exception as e:
            print(f'try make dir "{object_path}" failed, {e}.')
        object_db_file = os.path.join(object_path, f'{file_name}.db')
        object_cmd_file = os.path.join(object_path, f'{file_name}.txt')
        #
        lines_for_db = []
        lines_for_cmd = []
        for device_item in DeviceRegistered.values():
            lines_for_cmd.append(f'drvAsynIPPortConfigure("{device_item.name}", "{device_item.address}", 0, 0, 1)\n')
            lines_for_cmd.append(f'modbusInterposeConfig("{device_item.name}", 0, 2000, 0)\n')
        else:
            lines_for_cmd.append('\n')
        for pv_item in pl:
            pv_item.gen_prepare()
            lines_for_db.extend(pv_item.gen_db_lines())
            lines_for_cmd.extend(pv_item.gen_config_lines())
        else:
            with open(object_db_file, 'w') as f:
                f.writelines(lines_for_db)
            with open(object_cmd_file, 'w') as f:
                f.writelines(lines_for_cmd)
            print(DeviceRegistered)
