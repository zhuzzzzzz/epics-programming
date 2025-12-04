import os.path
import socket
import sys
from openpyxl import load_workbook

NameSeparator = ":"


class AsynDriver:
    def __init__(self):
        self.name = None
        self.device_name = None
        self.address = None
        self.info = None

        self.baud = None
        self.bits = None
        self.parity = None
        self.stop = None
        self.clocal = "N"
        self.crtscts = None

    def __str__(self):
        return f"{self.name}"

    def __repr__(self):
        return f"{self.name}"

    def is_ipv4_port(self):
        try:
            ip, port = self.address.rsplit(":", 1)
            socket.inet_aton(ip)
            port = int(port)
            return 0 <= port <= 65535
        except (ValueError, OSError, AttributeError):
            return False

    @property
    def config_lines(self):
        if self.is_ipv4_port():
            drvConfigLine = f'drvAsynIPPortConfigure("{self.name}", "{self.address}")\n'
        else:
            drvConfigLine = (
                f'drvAsynSerialPortConfigure("{self.name}", "{self.address}")\n'
            )
        drvConfigLine += f'asynSetOption("{self.name}", 0, "baud", "{self.baud}")\n'
        drvConfigLine += f'asynSetOption("{self.name}", 0, "bits", "{self.bits}")\n'
        drvConfigLine += f'asynSetOption("{self.name}", 0, "parity", "{self.parity}")\n'
        drvConfigLine += f'asynSetOption("{self.name}", 0, "stop", "{self.stop}")\n'
        drvConfigLine += f'asynSetOption("{self.name}", 0, "clocal", "{self.clocal}")\n'
        drvConfigLine += (
            f'asynSetOption("{self.name}", 0, "crtscts", "{self.crtscts}")\n'
        )
        return drvConfigLine

    @property
    def common_db_lines(self):
        return (
            f"# Error String Record\n"
            f"# 执行部分命令导致返回信息与接收预期不匹配时, 用此PV接收\n"
            f'record(stringin, "{self.name}:ErrorMessage"){{\n'
            f'\tfield(DESC, "Mismatch Error Message")\n'
            f"}}\n"
            f"\n"
            f"# Command Response Record\n"
            f"# 执行部分命令会返回执行情况状态码, 用此PV接收\n"
            f'record(stringin, "{self.name}:CommandResponse"){{\n'
            f'\tfield(DESC, "Executing Command Response")\n'
            f"}}\n"
            f"\n"
            f"# Raw Command Record\n"
            f"# 用此PV可以直接向设备发送串口命令并接受回复, 当发送的命令不产生回复时触发ReplyTimeout\n"
            f'record(stringout, "{self.name}:sendRawCommand"){{\n'
            f'\tfield(DESC, "Send Raw Command String to Device")\n'
            f'\tfield(DTYP, "stream")\n'
            f'\tfield(OUT, "@$(ProtocolFile) sendRawCommand({self.name}:ErrorMessage {self.name}:recvResponse) $(SerialPort)")\n'
            f"}}\n"
            f'record(lsi, "{self.name}:recvResponse"){{\n'
            f'\tfield(DESC, "Receive Raw Command Response")\n'
            f'\tfield(SIZV, "256")\n'
            f"}}\n"
            f"\n"
        )

    @property
    def common_proto_lines(self):
        return (
            f"# RawCommand\n"
            f"sendRawCommand{{\n"
            f'\tout "%s";\n'
            f'\tin "%(\$2)s";\n'
            f'\t@mismatch{{ in "%(\$1)39c"; disconnect; }}\n'
            f"\t@replytimeout{{ disconnect; }}\n"
            f"}}\n"
            f"\n"
        )


DriverRegistered = {}


class StreamData:
    def __init__(self):
        self.name = None
        self.asyn_driver = None
        self.data_operation = None

        self.terminator = None
        self.in_terminator = None
        self.out_terminator = None
        self.out_command = None
        self.in_match_str = None

        self.intr = None
        self.with_init = None

    @property
    def proto_name(self):
        return f"get{self.name}" if self.data_operation == "r" else f"set{self.name}"

    @property
    def protocol(self):
        line = f"{self.proto_name}{{\n"
        #
        terminator_string = self.get_terminator_string()
        if terminator_string:
            line += terminator_string
        #
        if self.intr:
            line += f'\tin "{self.in_match_str}";\n'
        else:
            line += f'\tout "{self.out_command}";\n' if self.out_command else ""
            line += f'\tin "{self.in_match_str}";\n' if self.in_match_str else ""
            line += f'\t@mismatch{{ in "%(\$1)39c"; disconnect; }}\n'
            line += f"\t@replytimeout{{ disconnect; }}\n"
        #
        if self.with_init and self.data_operation == "w":
            get_name = f"{self.asyn_driver}-{self.name}-r"
            if get_name in StreamDataRegistered:
                get_stream_data = StreamDataRegistered[get_name]
                if not get_stream_data.intr:
                    line += f"\t@init{{ get{self.name}; }}\n"
                else:
                    line += "\t@init{\n"
                    line += (
                        f'\t\tout "{get_stream_data.out_command}";\n'
                        if get_stream_data.out_command
                        else ""
                    )
                    line += (
                        f'\t\tin "{get_stream_data.in_match_str}";\n'
                        if get_stream_data.in_match_str
                        else ""
                    )
                    line += f'\t\t@mismatch{{ in "%(\$1)39c"; disconnect; }}\n'
                    line += f"\t\t@replytimeout{{ disconnect; }}\n"
                    line += "\t}\n"
        line += "}\n"
        return line

    def __str__(self):
        return f"{self.asyn_driver}-{self.name}-{self.data_operation}"

    def __repr__(self):
        return f"{self.asyn_driver}-{self.name}-{self.data_operation}"

    @staticmethod
    def get_global_terminator_settings():
        for drv in DriverRegistered.values():
            terminator_count_dict = {}
            in_terminator_count_dict = {}
            out_terminator_count_dict = {}
            for sdata in StreamDataRegistered.values():
                if sdata.asyn_driver.name != drv.name:
                    continue
                if sdata.terminator:
                    if sdata.terminator not in terminator_count_dict:
                        terminator_count_dict[sdata.terminator] = 1
                    else:
                        terminator_count_dict[sdata.terminator] += 1
                if sdata.in_terminator:
                    if sdata.in_terminator not in in_terminator_count_dict:
                        in_terminator_count_dict[sdata.in_terminator] = 1
                    else:
                        in_terminator_count_dict[sdata.in_terminator] += 1
                if sdata.out_terminator:
                    if sdata.out_terminator not in out_terminator_count_dict:
                        out_terminator_count_dict[sdata.out_terminator] = 1
                    else:
                        out_terminator_count_dict[sdata.out_terminator] += 1
            if terminator_count_dict:
                max_key = max(terminator_count_dict, key=terminator_count_dict.get)
                DrvCommonTerminator[drv.name] = max_key
            if in_terminator_count_dict:
                max_key = max(
                    in_terminator_count_dict, key=in_terminator_count_dict.get
                )
                DrvCommonInTerminator[drv.name] = max_key
            if out_terminator_count_dict:
                max_key = max(
                    out_terminator_count_dict, key=out_terminator_count_dict.get
                )
                DrvCommonOutTerminator[drv.name] = max_key

    def get_terminator_string(self):
        res = ""
        drv_name = self.asyn_driver.name
        if self.terminator:
            if (
                drv_name in DrvCommonTerminator
                and self.terminator != DrvCommonTerminator[drv_name]
            ):
                res += f"\tTerminator = {self.terminator};\n"
        if self.in_terminator:
            if (
                drv_name in DrvCommonInTerminator
                and self.in_terminator != DrvCommonInTerminator[drv_name]
            ):
                res += f"\tInTerminator = {self.in_terminator};\n"
        if self.out_terminator:
            if (
                drv_name in DrvCommonOutTerminator
                and self.out_terminator != DrvCommonOutTerminator[drv_name]
            ):
                res += f"\tOutTerminator = {self.out_terminator};\n"
        return res


StreamDataRegistered = {}

DrvCommonTerminator = {}
DrvCommonInTerminator = {}
DrvCommonOutTerminator = {}


class StreamDeviceRecord:
    def __init__(self, name="default", separator=NameSeparator):
        self.line_number = None
        self.name = name
        self.name_prefix = None
        self.name_separator = separator
        self.type = None  # 'ai' 'ao' 'bi' 'bo'
        self.scan = None
        self.desc = None
        self.prec = None
        self.egu = None
        self.other_fields = []

        self.stream_data = None

    def __str__(self):
        self.gen_prepare()
        return self.gen_db_lines() + self.gen_proto_lines()

    def __repr__(self):
        return f'"PV {self.name_prefix + self.name_separator + self.name}"'

    @staticmethod
    def get_field_line(field_str):
        return f"field({field_str})"

    def gen_prepare(self):
        #
        pass

    def gen_proto_lines(self):
        return self.stream_data.protocol

    def gen_db_lines(self):
        db_lines = [
            f'record({self.type}, "{self.name_prefix + self.name_separator + self.name}"){{\n',
        ]
        # DTYP
        line = self.get_field_line(f'DTYP, "stream"')
        db_lines.append(f"\t{line}\n")
        # INP or OUT
        if "o" in self.type.lower():
            if self.stream_data.in_match_str:
                line = self.get_field_line(
                    f"OUT, "
                    f'"@$(ProtocolFile) {self.stream_data.proto_name}('
                    f"{self.stream_data.asyn_driver.name}:ErrorMessage {self.stream_data.asyn_driver.name}:CommandResponse) "
                    f'$(SerialPort)"'
                )
            else:
                line = self.get_field_line(
                    f"OUT, "
                    f'"@$(ProtocolFile) {self.stream_data.proto_name}('
                    f"{self.stream_data.asyn_driver.name}:ErrorMessage) "
                    f'$(SerialPort)"'
                )
        else:
            line = self.get_field_line(
                f"INP, "
                f'"@$(ProtocolFile) {self.stream_data.proto_name}('
                f"{self.stream_data.asyn_driver.name}:ErrorMessage) "
                f'$(SerialPort)"'
            )
        db_lines.append(f"\t{line}\n")
        # SCAN
        if self.scan:
            line = self.get_field_line(f'SCAN, "{self.scan}"')
            db_lines.append(f"\t{line}\n")
        # DESC
        if self.desc:
            line = self.get_field_line(f'DESC, "{self.desc}"')
            db_lines.append(f"\t{line}\n")
        # PREC
        if self.prec:
            line = self.get_field_line(f'PREC, "{self.prec}"')
            db_lines.append(f"\t{line}\n")
        # EGU
        if self.egu:
            line = self.get_field_line(f'EGU, "{self.egu}"')
            db_lines.append(f"\t{line}\n")
        # other other_fields
        if self.other_fields:
            for item in self.other_fields:
                if item:
                    line = self.get_field_line(item)
                    db_lines.append(f"\t{line}\n")
        # end
        db_lines.append(f"}}\n")
        return "".join(db_lines)


# return the EXCEL table contents, merged_cells are handled.
def get_excel_list(
    file_path="./StreamDevice2db.xlsx", sheet_name="WorkArea", verbosity=0
):
    workbook = load_workbook(file_path)
    sheet_loaded = workbook[sheet_name]
    # print(sheet_loaded.calculate_dimension())

    excel_list = []
    for row in sheet_loaded.iter_rows(values_only=True):
        excel_list.append(list(row))

    for merged_cell in sheet_loaded.merged_cells:
        merged_value = sheet_loaded.cell(
            row=merged_cell.min_row, column=merged_cell.min_col
        ).value
        for i in range(merged_cell.bounds[0] - 1, merged_cell.bounds[2]):
            for j in range(merged_cell.bounds[1] - 1, merged_cell.bounds[3]):
                excel_list[j][i] = merged_value

    excel_list_not_empty_row = []
    for line in excel_list:
        if any(line):
            excel_list_not_empty_row.append(line)
    excel_list = excel_list_not_empty_row

    if verbosity >= 2:
        for line_list in excel_list:
            for item in line_list:
                print(item, end=" ")
            else:
                print("")

    return excel_list


# fill pv with EXCEL data, return the dict of device and pv list
# {'device_name': 'pv_list'}
def handle_excel_list(excel_list):
    pv_title = excel_list[0]
    driver_pvs = {}
    for j in range(1, len(excel_list)):
        line_number = j + 1
        #
        pv_temp = StreamDeviceRecord()
        pv_temp.line_number = line_number
        #
        driver_temp = AsynDriver()
        #
        sdata_temp = StreamData()
        #
        for i in range(len(pv_title)):
            if not pv_title[i]:
                continue
            # print(type(excel_list[j][i]))
            if "设备名称" in pv_title[i]:
                device_name = excel_list[j][i]
                if not device_name:
                    device_name = "UnknownDevice"
                    print(
                        f"警告: 表格第{line_number}行未配置StreamDevice设备名称, 将使用默认名称UnknownDevice"
                    )
                driver_temp.name = device_name
                driver_temp.device_name = device_name
            elif "设备地址" in pv_title[i]:
                device_address = excel_list[j][i]
                if not device_address:
                    print(f"错误: 表格第{line_number}行未配置StreamDevice设备地址")
                    exit(1)
                driver_temp.address = device_address
            elif "设备信息" in pv_title[i]:
                device_info = excel_list[j][i]
                driver_temp.info = device_info
            elif "波特率" in pv_title[i]:
                baud = excel_list[j][i]
                if not isinstance(baud, int):
                    print(f"警告: 表格第{line_number}行未配置正确的波特率")
                    exit(1)
                driver_temp.baud = baud
            elif "比特位数" in pv_title[i]:
                bits = excel_list[j][i]
                if not isinstance(bits, int):
                    print(f"警告: 表格第{line_number}行未配置正确的比特位数")
                    exit(1)
                driver_temp.bits = bits
            elif "奇偶校验" in pv_title[i]:
                parity = excel_list[j][i]
                if parity not in ("none", "odd", "even", "mark", "space"):
                    print(f"错误: 表格第{line_number}行未配置正确的奇偶校验")
                    exit(1)
                driver_temp.parity = parity
            elif "停止位" in pv_title[i]:
                stop = excel_list[j][i]
                if not isinstance(stop, int):
                    print(f"错误: 表格第{line_number}行未配置正确的停止位")
                    exit(1)
                driver_temp.stop = stop
            elif "启用RTS/CTS" in pv_title[i]:
                crtscts = excel_list[j][i]
                if crtscts not in ("是", "否"):
                    print(f"错误: 表格第{line_number}行未配置正确的流量控制选项")
                    exit(1)
                if crtscts == "是":
                    crtscts = "Y"
                else:
                    crtscts = "N"
                driver_temp.crtscts = crtscts
            elif "输入终止符" in pv_title[i]:
                in_terminator = excel_list[j][i]
                if not in_terminator:
                    print(f"错误: 表格第{line_number}行未配置正确的输入终止符")
                    exit(1)
                sdata_temp.in_terminator = in_terminator
            elif "输出终止符" in pv_title[i]:
                out_terminator = excel_list[j][i]
                if not out_terminator:
                    print(f"错误: 表格第{line_number}行未配置正确的输出终止符")
                    exit(1)
                sdata_temp.out_terminator = out_terminator
            elif "终止符" in pv_title[i]:
                terminator = excel_list[j][i]
                if not terminator:
                    print(f"错误: 表格第{line_number}行未配置正确的终止符")
                    exit(1)
                sdata_temp.terminator = terminator
            elif "数据名称" in pv_title[i]:
                data_name = excel_list[j][i]
                if not data_name:
                    print(f"错误: 表格第{line_number}行未配置数据名称")
                    exit(1)
                sdata_temp.name = data_name
            elif "数据操作" in pv_title[i]:
                data_opertation = excel_list[j][i]
                if not data_opertation.lower() in ("r", "w"):
                    print(f"错误: 表格第{line_number}行未配置数据操作")
                    exit(1)
                sdata_temp.data_operation = data_opertation
            elif "输出命令字符串" in pv_title[i]:
                out_command = excel_list[j][i]
                sdata_temp.out_command = out_command
            elif "输入匹配字符串" in pv_title[i]:
                in_match_str = excel_list[j][i]
                sdata_temp.in_match_str = in_match_str
            elif "Record类型" in pv_title[i]:
                r_type = excel_list[j][i]
                if not r_type:
                    print(f"错误: 表格第{line_number}行未配置Record类型")
                    exit(1)
                pv_temp.type = r_type
                if "o" in r_type.lower():
                    sdata_temp.with_init = True
                else:
                    sdata_temp.with_init = False
            elif "PV前缀" in pv_title[i]:
                name_prefix = excel_list[j][i]
                if not name_prefix:
                    print(f"错误: 表格第{line_number}行未配置PV名称前缀")
                    exit(1)
                pv_temp.name_prefix = name_prefix
            elif "PV后缀" in pv_title[i]:
                name = excel_list[j][i]
                if not name:
                    print(f"错误: 表格第{line_number}行未配置PV名称后缀")
                    exit(1)
                pv_temp.name = name
            elif "更新周期" in pv_title[i]:
                pv_temp.scan = excel_list[j][i]
                if pv_temp.scan == "I/O Intr":
                    sdata_temp.intr = True
                else:
                    sdata_temp.intr = False
            elif "PV描述" in pv_title[i]:
                pv_temp.desc = excel_list[j][i]
            elif "数据精度" in pv_title[i]:
                pv_temp.prec = excel_list[j][i]
            elif "数据单位" in pv_title[i]:
                pv_temp.egu = excel_list[j][i]
            elif "其他EPICS字段" in pv_title[i]:
                pv_temp.other_fields.append(excel_list[j][i])
        #
        if str(driver_temp) not in DriverRegistered.keys():
            DriverRegistered[str(driver_temp)] = driver_temp
            driver_pvs[driver_temp.name] = []
        sdata_temp.asyn_driver = DriverRegistered[str(driver_temp)]
        if str(sdata_temp) not in StreamDataRegistered.keys():
            StreamDataRegistered[str(sdata_temp)] = sdata_temp
        pv_temp.stream_data = sdata_temp
        #
        driver_pvs[device_name].append(pv_temp)
    StreamData.get_global_terminator_settings()
    return driver_pvs


if __name__ == "__main__":
    debug_level = 0
    for item in sys.argv:
        if "-v" in item.lower():
            sys.argv.remove(item)
            debug_level = item.count("v")
            break
    if debug_level >= 1:
        print(" ".join(sys.argv), "\t", f"debug_level={debug_level}")
    if len(sys.argv) < 2 or len(sys.argv) > 3:
        print(
            f"Usage:\n"
            f'python {os.path.basename(sys.argv[0])} "excel file"\n'
            f'python {os.path.basename(sys.argv[0])} "excel file" "sheet name"\n'
        )
        exit(1)
    #
    if not os.path.isfile(sys.argv[1]):
        print(f"{sys.argv[0]}: Invalid excel path: {sys.argv[1]}!")
        exit(1)
    #
    driver_pv_list_dict = None
    if len(sys.argv) == 2:
        driver_pv_list_dict = handle_excel_list(
            get_excel_list(file_path=sys.argv[1], verbosity=debug_level)
        )
    else:  # len(sys.argv) == 3
        driver_pv_list_dict = handle_excel_list(
            get_excel_list(
                file_path=sys.argv[1], sheet_name=sys.argv[2], verbosity=debug_level
            )
        )
    #
    object_path = os.path.join(os.getcwd(), "res")
    os.makedirs(object_path, exist_ok=True)
    #
    for drv in DriverRegistered.values():
        file_name = drv.name
        object_db_file = os.path.join(object_path, f"{file_name}.db")
        object_cmd_file = os.path.join(object_path, f"{file_name}.cmd")
        object_protocol_file = os.path.join(object_path, f"{file_name}.proto")
        #
        lines_for_db = []
        lines_for_proto = []
        #
        LF_flag = False
        if drv.name in DrvCommonTerminator:
            lines_for_proto.append(f"Terminator = {DrvCommonTerminator[drv.name]};\n")
            LF_flag = True
        if drv.name in DrvCommonInTerminator:
            lines_for_proto.append(
                f"InTerminator = {DrvCommonInTerminator[drv.name]};\n"
            )
            LF_flag = True
        if drv.name in DrvCommonOutTerminator:
            lines_for_proto.append(
                f"OutTerminator = {DrvCommonOutTerminator[drv.name]};\n"
            )
            LF_flag = True
        if LF_flag:
            lines_for_proto.append("\n")
        lines_for_proto.append(drv.common_proto_lines)
        #
        for pv_item in driver_pv_list_dict[drv.name]:
            pv_item.gen_prepare()
            lines_for_db.extend(pv_item.gen_db_lines())
            lines_for_proto.append(pv_item.gen_proto_lines())
        else:
            with open(object_db_file, "w") as f:
                f.writelines(drv.common_db_lines)
                f.writelines(lines_for_db)
                print(f'写入db文件: "{object_db_file}"')
            with open(object_protocol_file, "w") as f:
                f.writelines(lines_for_proto)
                print(f'写入protocol文件: "{object_protocol_file}"')
        #
        with open(object_cmd_file, "w") as f:
            f.writelines(drv.config_lines)
            print(f'写入cmd文件: "{object_cmd_file}"')
    if debug_level >= 1:
        print("DriverRegistered:", DriverRegistered)
        print("StreamDataRegistered", StreamDataRegistered)
        print("DrvCommonTerminator", DrvCommonTerminator)
        print("DrvCommonInTerminator", DrvCommonInTerminator)
        print("DrvCommonOutTerminator", DrvCommonOutTerminator)
        print("driver_pv_list_dict", driver_pv_list_dict)
