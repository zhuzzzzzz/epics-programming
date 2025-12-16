import os.path
import sys
from openpyxl import load_workbook

NameSeparator = ':'

drvUser2DTYP = {
    'BCD': 'asynInt32',
    'UINT16': 'asynUInt32Digital',
    'INT16': 'asynInt32',
    'UINT32': 'asynInt32',
    'INT32': 'asynInt32',
    'UINT64': 'asynInt64',
    'INT64': 'asynInt64',
    'FLOAT32': 'asynFloat64',
    'FLOAT64': 'asynFloat64',
    'STRING': 'asynInt32',
    'ZSTRING': 'asynInt32',
}

SupportedDataFormat = {
    'BCD': ['_UNSIGNED', '_SIGNED'],
    'INT16': ['', 'SM'],
    'UINT16': [''],
    'INT32': ['_LE', '_LE_BS', '_BE', '_BE_BS'],
    'UINT32': ['_LE', '_LE_BS', '_BE', '_BE_BS'],
    'INT64': ['_LE', '_LE_BS', '_BE', '_BE_BS'],
    'UINT64': ['_LE', '_LE_BS', '_BE', '_BE_BS'],
    'FLOAT32': ['_LE', '_LE_BS', '_BE', '_BE_BS'],
    'FLOAT64': ['_LE', '_LE_BS', '_BE', '_BE_BS'],
    'STRING': ['_HIGH', '_LOW', '_HIGH_LOW', '_LOW_HIGH'],
    'ZSTRING': ['_HIGH', '_LOW', '_HIGH_LOW', '_LOW_HIGH'],
}
SupportedDataFormatList = []
for key, suffixes in SupportedDataFormat.items():
    for suffix in suffixes:
        SupportedDataFormatList.append(key + suffix)


class ModbusDevice:
    def __init__(self):
        self.name = None
        self.device_type = None
        self.address = None  # ip:port
        self.info = None

    def __str__(self):
        return f'(名称: "{self.name}", 类型: "{self.device_type}", 地址: "{self.address}", 信息: "{self.info}")'

    def __repr__(self):
        return f'(名称: "{self.name}", 类型: "{self.device_type}", 地址: "{self.address}", 信息: "{self.info}")'


DeviceRegistered = {}


class ModbusDriver:
    def __init__(self):
        self.name = None
        self.device = None
        self.slave_id = None
        self.function_code = None
        self.memory_address = None
        self.memory_length = None

    def __str__(self):
        return f'(名称: "{self.name}", 类型: "Modbus Driver")'

    def __repr__(self):
        return f'(名称: "{self.name}", 类型: "Modbus Driver")'

    @property
    def config_line(self):
        # drvModbusAsynConfigure(portName,
        #                        tcpPortName,
        #                        slaveAddress,
        #                        modbusFunction,
        #                        modbusStartAddress,
        #                        modbusLength,
        #                        dataType,
        #                        pollMsec,
        #                        plcType);
        return (f'drvModbusAsynConfigure('
                f'"{self.name}", '
                f'"{self.device}", '
                f'{self.slave_id}, '
                f'{self.function_code}, '
                f'{self.memory_address}, '
                f'{self.memory_length}, '
                f'0, '
                f'100, '
                f'"{self.device}")\n')


DriverRegistered = {}


class ModbusRecord:
    def __init__(self, name='default', seperator=NameSeparator):
        self.line_number = None
        self.name = name
        self.name_prefix = None
        self.name_seperator = seperator
        self.type = None  # 'ai' 'ao' 'bi' 'bo'
        self.scan = None
        self.desc = None
        self.prec = None
        self.egu = None
        self.other_fields = []
        self.device = None
        self.slave_id = None
        self.memory_address = None  # memory address for data in PLC
        self.memory_offset = None
        self.data_length = None  # data length
        self.memory_length = None  # driver memory length
        self.device_access = None  # 'r', 'w'
        self.drvUser_prefix = None
        self.drvUser_suffix = None
        self.modbus_funcode = None  # 1, 2, 3, 4, 5, 6, 15, 16
        self.interface_name = None
        self.memory_address_mask = None

    def __str__(self):
        self.gen_prepare()
        return self.gen_config_line() + self.gen_db_lines()

    def __repr__(self):
        return f'"PV {self.name_prefix + self.name_seperator + self.name}"'

    @staticmethod
    def get_field_line(field_str):
        return f'field({field_str})'

    def gen_prepare(self):
        # 非必填字段
        if not self.line_number:
            self.line_number = 'unknown'
        if not self.name_prefix:
            self.name_prefix = ''
        if not self.name:
            self.name = ''
        if not self.drvUser_prefix:
            self.drvUser_prefix = ''
        if not self.drvUser_suffix:
            self.drvUser_suffix = ''
        # 必填字段
        if not self.device:
            print(f'错误: 表格第{self.line_number}行未配置PLC设备')
            exit(1)
        if not self.slave_id and not isinstance(self.slave_id, int):
            print(f'错误: 表格第{self.line_number}行未配置正确的PLC设备从站号')
            exit(1)
        # if not isinstance(self.memory_address, int) or self.memory_address < 0:
        #     print(f'错误: 表格第{self.line_number}行未配置正确的PV对应的设备地址')
        #     exit(1)
        if not self.memory_offset and not isinstance(self.memory_offset, int):
            print(f'错误: 表格第{self.line_number}行未配置正确的PV对应的设备地址偏移量')
            exit(1)
        if not self.data_length or not isinstance(self.data_length, int):
            print(f'错误: 表格第{self.line_number}行未配置正确的PV数据长度')
            exit(1)
        if not self.device_access in ('r', 'w'):
            print(f'错误: 表格第{self.line_number}行未配置正确的PV访问方式')
            exit(1)
        if self.modbus_funcode is not None and not self.modbus_funcode in (1, 2, 3, 4, 5, 6, 15, 16):
            print(f'错误: 表格第{self.line_number}行未配置正确的Modbus功能码')
            exit(1)
        # 数据校验
        if self.drvUser_prefix + self.drvUser_suffix not in SupportedDataFormatList:
            print(f'错误: 表格第{self.line_number}行未配置正确的PV数据格式')
            exit(1)
        #
        if self.modbus_funcode is None:
            if self.device_access == 'r':
                self.modbus_funcode = 3
                if self.memory_address_mask:
                    self.type = 'bi'
                else:
                    self.type = 'ai'
            else:
                self.modbus_funcode = 16
                if self.memory_address_mask:
                    self.type = 'bo'
                else:
                    self.type = 'ao'
        elif self.modbus_funcode in (1, 2):
            if self.device_access != 'r':
                print(f'错误: 表格第{self.line_number}行配置的读写策略与Modbus功能码不匹配')
                exit(1)
            self.memory_address_mask = '0x1'
            self.type = 'bi'
        elif self.modbus_funcode in (5, 15):
            if self.device_access != 'w':
                print(f'错误: 表格第{self.line_number}行配置的读写策略与Modbus功能码不匹配')
                exit(1)
            self.memory_address_mask = '0x1'
            self.type = 'bo'
        elif self.modbus_funcode in (3, 4):
            if self.device_access != 'r':
                print(f'错误: 表格第{self.line_number}行配置的读写策略与Modbus功能码不匹配')
                exit(1)
            if self.memory_address_mask:
                self.type = 'bi'
            else:
                self.type = 'ai'
        elif self.modbus_funcode in (6, 16):
            if self.device_access != 'w':
                print(f'错误: 表格第{self.line_number}行配置的读写策略与Modbus功能码不匹配')
                exit(1)
            if self.memory_address_mask:
                self.type = 'bo'
            else:
                self.type = 'ao'
        else:
            print(f'错误: 表格第{self.line_number}行配置了不支持的Modbus功能码: {self.modbus_funcode}')
            exit(1)
        #
        if self.modbus_funcode in (1, 2, 5, 15):
            self.memory_length = self.memory_offset + self.data_length
        else:  # self.modbus_funcode in (3, 4, 6, 16):
            import math
            self.memory_length = self.memory_offset + math.ceil(self.data_length / 2)
        #
        self.interface_name = f'{self.device}_{self.slave_id}_Addr{self.memory_address}_FC{self.modbus_funcode}'

    def gen_config_line(self):
        if self.interface_name not in DriverRegistered:
            modbus_driver = ModbusDriver()
            modbus_driver.name = self.interface_name
            modbus_driver.device = self.device
            modbus_driver.slave_id = self.slave_id
            modbus_driver.function_code = self.modbus_funcode
            modbus_driver.memory_address = self.memory_address
            modbus_driver.memory_length = self.memory_length
            DriverRegistered[self.interface_name] = modbus_driver
        else:
            if self.memory_length > DriverRegistered[self.interface_name].memory_length:
                DriverRegistered[self.interface_name].memory_length = self.memory_length
        return DriverRegistered[self.interface_name].config_line

    def gen_db_lines(self):
        db_lines = [
            f'record({self.type}, "{self.name_prefix + self.name_seperator + self.name}"){{\n',
        ]
        # DTYP
        line = self.get_field_line(f'DTYP, "{drvUser2DTYP[self.drvUser_prefix]}"')
        db_lines.append(f'\t{line}\n')
        # INP or OUT
        if self.device_access == 'r' and 'a' in self.type:
            line = self.get_field_line(
                f'INP, "@asyn({self.interface_name} {self.memory_offset}){self.drvUser_prefix + self.drvUser_suffix}"')
        elif self.device_access == 'r' and 'b' in self.type:
            line = self.get_field_line(
                f'INP, "@asynMask({self.interface_name} {self.memory_offset} {self.memory_address_mask}){self.drvUser_prefix + self.drvUser_suffix}"')
        elif self.device_access == 'w' and 'a' in self.type:
            line = self.get_field_line(
                f'OUT, "@asyn({self.interface_name} {self.memory_offset}){self.drvUser_prefix + self.drvUser_suffix}"')
        elif self.device_access == 'w' and 'b' in self.type:
            line = self.get_field_line(
                f'OUT, "@asynMask({self.interface_name} {self.memory_offset} {self.memory_address_mask}){self.drvUser_prefix + self.drvUser_suffix}"')
        else:
            print(f'{self.name}.gen_db_lines failed for line "INP" or "OUT". {self.device_access} {self.type}')
        db_lines.append(f'\t{line}\n')
        # SCAN
        if self.scan:
            line = self.get_field_line(f'SCAN, "{self.scan}"')
            db_lines.append(f'\t{line}\n')
        # DESC
        if self.desc:
            line = self.get_field_line(f'DESC, "{self.desc}"')
            db_lines.append(f'\t{line}\n')
        # PREC
        if self.prec:
            line = self.get_field_line(f'PREC, "{self.prec}"')
            db_lines.append(f'\t{line}\n')
        # EGU
        if self.egu:
            line = self.get_field_line(f'EGU, "{self.egu}"')
            db_lines.append(f'\t{line}\n')
        # other other_fields
        if self.other_fields:
            for item in self.other_fields:
                if item:
                    line = self.get_field_line(item)
                    db_lines.append(f'\t{line}\n')
        # end
        db_lines.append(f'}}\n')
        return ''.join(db_lines)


# return the EXCEL table contents, merged_cells are handled.
def get_excel_list(file_path='./Modbus2db.xlsx', sheet_name='WorkArea', verbosity=0):
    workbook = load_workbook(file_path)
    sheet_loaded = workbook[sheet_name]
    # print(sheet_loaded.calculate_dimension())

    excel_list = []
    for row in sheet_loaded.iter_rows(values_only=True):
        excel_list.append(list(row))

    for merged_cell in sheet_loaded.merged_cells:
        merged_value = sheet_loaded.cell(row=merged_cell.min_row, column=merged_cell.min_col).value
        for i in range(merged_cell.bounds[0] - 1, merged_cell.bounds[2]):
            for j in range(merged_cell.bounds[1] - 1, merged_cell.bounds[3]):
                excel_list[j][i] = merged_value

    excel_list_not_empty_row = []
    for line in excel_list:
        if any(line):
            excel_list_not_empty_row.append(line)
    excel_list = excel_list_not_empty_row

    if verbosity >= 2:
        for line_list in excel_list:
            for item in line_list:
                print(item, end=' ')
            else:
                print('')

    return excel_list


# fill pv with EXCEL data, return the dict of device and pv list
# {'device_name': 'pv_list'}
def handle_excel_list(excel_list):
    pv_title = excel_list[0]
    device_pvs = {}
    for j in range(1, len(excel_list)):
        line_number = j + 1
        pv_temp = ModbusRecord()
        pv_temp.line_number = line_number
        #
        device_name = None
        device_address = None
        device_info = None
        #
        for i in range(len(pv_title)):
            if not pv_title[i]:
                continue
            # print(type(excel_list[j][i]))
            if 'PLC名称' in pv_title[i]:
                device_name = excel_list[j][i]
                if not device_name:
                    device_name = 'UnknownDevice'
                    print(f'警告: 表格第{line_number}行未配置PLC设备名称, 将使用默认名称UnknownDevice')
                pv_temp.device = device_name
            elif 'IP:Port' in pv_title[i]:
                device_address = excel_list[j][i]
                if not device_address:
                    print(f'错误: 表格第{line_number}行未配置PLC设备地址')
                    exit(1)
            elif 'PLC信息' in pv_title[i]:
                device_info = excel_list[j][i]
            elif 'PLC从站号' in pv_title[i]:
                slave_id = excel_list[j][i]
                if not isinstance(slave_id, int):
                    print(f'警告: 表格第{line_number}行未配置正确的PLC设备从站号, 将使用默认从站号0')
                pv_temp.slave_id = slave_id
            elif 'Modbus功能码' in pv_title[i]:
                pv_temp.modbus_funcode = excel_list[j][i]
            elif 'Address' in pv_title[i]:
                memory_address = excel_list[j][i]
                if not memory_address or not isinstance(memory_address, int):
                    print(f'错误: 表格第{line_number}行未配置PV地址')
                    exit(1)
                pv_temp.memory_address = memory_address
            elif 'Offset' in pv_title[i]:
                memory_offset = excel_list[j][i]
                if not isinstance(memory_offset, int):
                    print(f'错误: 表格第{line_number}行未配置PV地址偏移, 无偏移请设置为0')
                    exit(1)
                pv_temp.memory_offset = memory_offset
            elif '数据操作' in pv_title[i]:
                device_access = excel_list[j][i]
                if device_access.lower() not in ('r', 'w', 'rw'):
                    print(f'错误: 表格第{line_number}行未配置正确的PV数据操作类型')
                    exit(1)
                pv_temp.device_access = device_access
            elif '数据长度' in pv_title[i]:
                data_length = excel_list[j][i]
                if not isinstance(data_length, int) or data_length <= 0:
                    print(f'错误: 表格第{line_number}行未配置正确的PV数据长度')
                    exit(1)
                pv_temp.data_length = data_length
            elif 'PV前缀' in pv_title[i]:
                name_prefix = excel_list[j][i]
                if not name_prefix:
                    print(f'错误: 表格第{line_number}行未配置PV名称前缀')
                    exit(1)
                pv_temp.name_prefix = name_prefix
            elif 'PV后缀' in pv_title[i]:
                name = excel_list[j][i]
                if not name:
                    print(f'错误: 表格第{line_number}行未配置PV名称后缀')
                    exit(1)
                pv_temp.name = name
            elif '更新周期' in pv_title[i]:
                pv_temp.scan = excel_list[j][i]
            elif 'PV描述' in pv_title[i]:
                pv_temp.desc = excel_list[j][i]
            elif '数据精度' in pv_title[i]:
                pv_temp.prec = excel_list[j][i]
            elif '数据单位' in pv_title[i]:
                pv_temp.egu = excel_list[j][i]
            elif '数据类型' in pv_title[i]:
                drvUser_prefix = excel_list[j][i]
                if not drvUser_prefix:
                    print(f'错误: 表格第{line_number}行未配置PV数据类型')
                    exit(1)
                pv_temp.drvUser_prefix = drvUser_prefix
            elif '数据格式' in pv_title[i]:
                pv_temp.drvUser_suffix = excel_list[j][i]
            elif '掩码' in pv_title[i]:
                pv_temp.memory_address_mask = excel_list[j][i]
            elif '其他EPICS字段' in pv_title[i]:
                pv_temp.other_fields.append(excel_list[j][i])
        #
        if device_name not in DeviceRegistered.keys():
            md = ModbusDevice()
            DeviceRegistered[device_name] = md
            md.name = device_name
            md.address = device_address
            md.device_type = 'Modbus'
            md.info = device_info
            device_pvs[device_name] = []
        #
        device_pvs[device_name].append(pv_temp)
    return device_pvs


if __name__ == '__main__':
    debug_level = 0
    for item in sys.argv:
        if '-v' in item.lower():
            sys.argv.remove(item)
            debug_level = item.count('v')
            break
    if debug_level >= 1:
        print(' '.join(sys.argv), '\t', f'debug_level={debug_level}')
    if len(sys.argv) < 2 or len(sys.argv) > 3:
        print(f'Usage:\n'
              f'python {os.path.basename(sys.argv[0])} "excel file"\n'
              f'python {os.path.basename(sys.argv[0])} "excel file" "sheet name"\n')
        exit(1)
    if not os.path.isfile(sys.argv[1]):
        print(f'{sys.argv[0]}: Invalid excel path: {sys.argv[1]}!')
        exit(1)
    device_pv_list_dict = None
    if len(sys.argv) == 2:
        device_pv_list_dict = handle_excel_list(get_excel_list(file_path=sys.argv[1], verbosity=debug_level))
    else:
        device_pv_list_dict = handle_excel_list(
            get_excel_list(file_path=sys.argv[1], sheet_name=sys.argv[2], verbosity=debug_level))
    #
    object_path = os.path.join(os.getcwd(), 'res')
    os.makedirs(object_path, exist_ok=True)
    for dev in device_pv_list_dict:
        file_name = dev
        object_db_file = os.path.join(object_path, f'{file_name}.db')
        object_cmd_file = os.path.join(object_path, f'{file_name}.cmd')
        #
        lines_for_db = []
        lines_for_cmd = []
        #
        for device_item in DeviceRegistered.values():
            if device_item.name == dev:
                lines_for_cmd.append(f'drvAsynIPPortConfigure("{device_item.name}", "{device_item.address}", 0, 0, 1)\n')
                lines_for_cmd.append(f'modbusInterposeConfig("{device_item.name}", 0, 2000, 0)\n')
        else:
            lines_for_cmd.append('\n')
        #
        for pv_item in device_pv_list_dict[dev]:
            pv_item.gen_prepare()
            pv_item.gen_config_line()
            lines_for_db.extend(pv_item.gen_db_lines())
        else:
            with open(object_db_file, 'w') as f:
                f.writelines(lines_for_db)
                print(f'写入db文件: "{object_db_file}"')
        #
        for driver_item in DriverRegistered.values():
            if driver_item.device == dev:
                lines_for_cmd.append(driver_item.config_line)
        else:
            with open(object_cmd_file, 'w') as f:
                f.writelines(lines_for_cmd)
                print(f'写入cmd文件: "{object_cmd_file}"')
    if debug_level >= 1:
        print(device_pv_list_dict)
        print(DeviceRegistered)
        print(DriverRegistered)
